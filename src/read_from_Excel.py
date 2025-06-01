# -*- coding: utf-8 -*-
"""
@author: Matthias Grajewski, FH Aachen University of Applied Sciences
Luis Hasenauer, FH Aachen University of Applied Sciences

This module provides functions for reading the MCDA data from an Excel file.

This file is part of the pysmaa python package, available at https://github.com/mgrajewski/pysmaa .
"""

import warnings
import numpy as np
import openpyxl as oxl
from typing import Tuple
from numpy import ndarray


def read_perfmat_from_excel(file_name: str, sheet_name: str) -> Tuple[np.ndarray, np.ndarray, np.ndarray, np.ndarray,
                                                                      list, list]:
    """
    This function reads the raw performance matrix perfmat from an Excel file. It is assumed that the performance matrix
    is contained in a worksheet with name sheet_name. This sheet title is case-sensitive.
    Data regarding the performance matrix are indicated by keywords. If a keyword is provided, then it is required to
    be in somewhere in the first column of that worksheet. In the row below a keyword, starting in the second column,
    the function expects the names of the criteria. In the second column, starting two rows below the keyword, the
    function expects the names of the alternatives:

    .. math::
        \\begin{bmatrix}
            \\text{keyword} & \\text{ } & \\text{ } & \\text{ }\\\\
            \\text{criteria} & \\text{1st crit} & \\cdots & \\text{last crit}\\\\
            \\text{1st alt} & p & p & p\\\\
            \\vdots & p & p & p\\\\
            \\text{last alt} & p & p & p
        \\end{bmatrix}


    Where :math:`p` represents the performance matrix.

    There must be no empty columns between a keyword and the subsequent matrix of data.
    The following keywords exist:
    "Mean" (mandatory) indicates that the data below refer to the mean values of the performance matrix.
    "Type" (optional). If "Type" does not exist, the performance matrix is supposed to have fixed entries. If it exists,
    the data below "Type" indicate the type of entry in the performance matrix (F: fixed number, U: uniformly
    distributed, N: normally distributed, C: custom distribution to be provided by the user)
    "Param": If "Type" is present and there are entries other than F or C, "Param" is mandatory. The keyword indicates
    that the subsequent data refer to parameters of the distributions. In the case of a normal distribution, this is the
    variance, in the case of a uniform distribution, its range is [(1-c)*Mean, (1+c)*Mean], where c is the corresponding
    parameter. If "Type" is not present, the data below "Param" is ignored.
    "Benefit" (optional) indicates if the respective criterion is a benefit (value True) or cost (value False). Instead
    of 'False', '0', 'NO', 'F', 'FALSCH' and lowercase pendants are accepted, instead of True", one may use '1', 'YES',
    'T', 'TRUE', 'WAHR' or lowercase variants of these keywords. If the keyword "Benefit" does not exist, every
    criterion is assumed to be connected to a benefit, and a warning is issued.

    The function returns the performance matrix perfmat_mean, the type matrix perfmat_types (default value: 0), the
    parameter matrix perfmat_params (default value: NaN) and the benefit vector benefit (default value: True).

    Parameters
    ----------
    file_name : string
        name of the Excel file to read from
    sheet_name : string
        name of the Excel work sheet to read perfmat from

    Returns
    -------
    perfmat_mean : 2D-numpy-array
    perfmat_types : 2D-numpy-array
    perfmat_param : 2D-numpy-array
    benefit : 1D-numpy-array (bool)

    criteria : list containing the names of the criteria
    alternatives:  list containing the names of the alternatives
    """

    list_of_types = ['F', 'U', 'N', 'C']
    int_vals_of_types = [0, 1, 2, 10]

    benefit_keywords = ('1', 'YES', 'T', 'TRUE', 'WAHR')
    cost_keywords = ('0', 'NO', 'F', 'FALSE', 'FALSCH')

    workbook = oxl.load_workbook(filename=file_name, data_only=True)

    # maximal row to consider at all (The openpyxl-function worksheet.max_row is sometimes unreliable and may return
    # too large values.)
    max_row = 300

    # maximal column to consider at all (The openpyxl-function worksheet.max_column is sometimes unreliable and may
    # return too large values)
    max_column = 100

    row_keyword_mean = max_row
    row_keyword_type = max_row
    row_keyword_param = max_row
    row_keyword_benefit = max_row

    # open Excel worksheet with name "sheet_name'
    worksheet = workbook[sheet_name]

    # maximal column with non-empty cells (as worksheet.max_column is
    # unreliable, we safeguard it with the initial max_column). The same
    # applies to max_row
    max_column = max(max_column, worksheet.max_column)
    max_row = max(max_row, worksheet.max_row)

    # find keyword "Type" in first column
    keyword_type_found = False
    for irow in range(1, max_row):
        if worksheet.cell(row=irow, column=1).value == 'Type':
            row_keyword_type = irow
            keyword_type_found = True
            break

    # find keyword "Param" in first column
    keyword_param_found = False
    for irow in range(1, max_row):
        if worksheet.cell(row=irow, column=1).value == 'Param':
            row_keyword_param = irow
            keyword_param_found = True
            break

    # find keyword "Mean" in first column
    keyword_mean_found = False
    for irow in range(1, max_row):
        if worksheet.cell(row=irow, column=1).value == 'Mean':
            row_keyword_mean = irow
            keyword_mean_found = True
            break

    # keyword 'Mean' is mandatory
    if not keyword_mean_found:
        err_msg = 'Keyword \'Mean\' was not found. Please provide a performance matrix following the keyword \'Mean\'.'
        raise NameError(err_msg)

    # find keyword "Benefit" in first column
    keyword_benefit_found = False
    for irow in range(1, max_row):
        if worksheet.cell(row=irow, column=1).value == 'Benefit':
            row_keyword_benefit = irow
            keyword_benefit_found = True
            break

    istart = row_keyword_mean
    istop = max_row

    # read the names of the criteria in the row below the keyword 'Mean'
    row_read = row_keyword_mean

    # find number of alternatives by considering the matrix following the keyword 'Mean'
    nalts = 0
    for irow in range(istart + 2, istop):
        if worksheet.cell(row=irow, column=1).value is not None:
            nalts = nalts + 1
        else:
            break

    # find number of criteria (first guess)
    ncrit = max_column - 1

    criteria = []
    alternatives = []

    # read in the first column after the keyword 'Mean' (contains the names of the criteria)
    for icol in range(2, ncrit + 2):
        if worksheet.cell(row=1 + row_read, column=icol).value is not None:
            criteria.append(worksheet.cell(row=1 + row_read, column=icol).value)

    # this is the real number of criteria (it could happen that there are some filled Excel cells apart from the
    # actual data such that max_column-1 is greater than the number of criteria)
    ncrit = len(criteria)

    # allocate matrices and initialise them with zeros
    perfmat_mean = np.zeros((nalts, ncrit))
    perfmat_type = np.zeros((nalts, ncrit), dtype=int)
    perfmat_param = np.full((nalts, ncrit), np.nan)
    benefit = np.ones(ncrit, dtype=bool)

    # read the names of the alternatives
    for irow in range(row_read + 2, nalts + 3):
        if worksheet.cell(row=irow, column=1).value is not None:
            alternatives.append(worksheet.cell(row=irow, column=1).value)
        else:
            err_msg = 'Reading the names of the alternatives failed due ' + \
                      'to empty cells. Please name every alternative.'
            raise NameError(err_msg)

    # read mean values of the performance matrix
    for irow in range(0, nalts):
        for icol in range(0, ncrit):
            if worksheet.cell(row=irow + 3, column=icol + 2).value is not None:
                perfmat_mean[irow, icol] = worksheet.cell(row=irow + 3, column=icol + 2).value
            else:
                err_msg = 'Reading the mean values for perfmat failed ' + \
                          'due to empty cells (row ' + str(irow + 3) + \
                          ', column ' + str(icol + 2) + ').'
                perfmat_mean[irow, icol] = np.nan
                raise NameError(err_msg)

    # if type info is provided, read it
    if keyword_type_found:

        criteria_type = []
        row_read = row_keyword_type
        # read in the first column after the keyword 'Type' (contains the names of the criteria)
        for icol in range(2, ncrit + 2):
            if worksheet.cell(row=1 + row_read, column=icol).value is not None:
                criteria_type.append(worksheet.cell(row=1 + row_read, column=icol).value)

        __check_criteria(criteria_type, criteria, 'Type')

        for irow in range(0, nalts):
            for j in range(0, ncrit):
                if worksheet.cell(row=irow + 2 + row_keyword_type, column=j + 2).value is None:
                    err_msg = 'Reading the type information of the performance matrix ' + \
                              'failed due to an empty cell (row ' + \
                              str(irow + 2 + row_keyword_type) + ', column ' + str(j + 2) + ').'
                    raise NameError(err_msg)

                aux = worksheet.cell(row=irow + 2 + row_keyword_type, column=j + 2).value

                if not isinstance(aux, str) or aux not in list_of_types:
                    err_msg = 'Reading the type information of the performance matrix failed due to ' + \
                              'invalid content in a cell (row ' + \
                              str(irow + 2 + row_keyword_type) + ', column ' + str(j + 2) + '). \n' + \
                              'A cell may contain only ' + str(list_of_types) + '.'
                    raise NameError(err_msg)

                perfmat_type[irow, j] = int_vals_of_types[list_of_types.index(aux)]

    # if there are entries in the type matrix being neither 'F' for a fixed number or 'C' for a custom distribution and
    # there is no parameter matrix at all: throw error message
    if np.any(np.logical_and(perfmat_type != int_vals_of_types[list_of_types.index('F')],
                             perfmat_type != int_vals_of_types[list_of_types.index('C')])) \
            and not keyword_param_found:
        err_msg = ('The type information of the performance matrix shows that there are some entries intended to be'
                   ' parametric distributions. \n Please provide a parameter matrix (keyword \'Param\').')
        raise NameError(err_msg)

    # if parameters for distributions are provided, read them
    if keyword_param_found:

        criteria_param = []
        row_read = row_keyword_param
        # read in the first column after the keyword 'Param' (contains the names of the criteria)
        for icol in range(2, ncrit + 2):
            if worksheet.cell(row=1 + row_read, column=icol).value is not None:
                criteria_param.append(worksheet.cell(row=1 + row_read, column=icol).value)

        __check_criteria(criteria_param, criteria, 'Type')

        for i in range(0, nalts):
            for j in range(0, ncrit):
                # read values, if present
                if worksheet.cell(row=i + 2 + row_keyword_param, column=j + 2).value is not None:
                    perfmat_param[i, j] = worksheet.cell(row=i + 2 + row_keyword_param, column=j + 2).value
                else:

                    #
                    if perfmat_type[i, j] not in [int_vals_of_types[list_of_types.index('F')],
                                                  int_vals_of_types[list_of_types.index('C')]]:
                        err_msg = 'Reading the parameter values of the perfomance matrix ' + \
                                  'failed due to an empty cell (row ' + \
                                  str(i + 2 + row_keyword_param) + ', column ' + str(j + 2) + ').'
                        raise NameError(err_msg)

    if keyword_benefit_found:
        for icol in range(0, ncrit):
            # read values, if present
            if worksheet.cell(row=2 + row_keyword_benefit, column=icol + 2).value is not None:
                cell_value = str(worksheet.cell(row=2 + row_keyword_benefit, column=icol + 2).value).upper()
                if cell_value in benefit_keywords:
                    benefit[icol] = True
                elif cell_value in cost_keywords:
                    benefit[icol] = False
                else:
                    err_msg = ('Invalid keyword \'' + cell_value + '\' for cost/benefit. Please use some of ' +
                               str(benefit_keywords) + ' or ' + str(cost_keywords) +
                               ' (keywords are not case-sensitive).')
                    raise NameError(err_msg)
            else:
                err_msg = 'No information on cost/benefit provided for criterion ' + criteria[icol] + '.'
                raise NameError(err_msg)
    else:
        warn_msg = 'No information on cost/benefit provided. We assume any criterion is associated to a benefit.'
        warnings.warn(warn_msg)

    # close the Excel file
    workbook.close()

    return perfmat_mean, perfmat_type, perfmat_param, benefit, criteria, alternatives


def read_rankings_from_excel(file_name: str) -> Tuple[list, list]:
    """
    This function reads the desired rankings from an Excel file.
    It is assumed that the data are contained in a worksheet called "Ranking". This sheet title is case-sensitive. The
    first row is reserved for the names of the stakeholders associated with the rankings, which are given column-wise
    in Excel. The rankings, which are sorted in descending order, are expected to start in the first column (highest
    cell is the preferred alternative).

    Parameters
    ----------
    file_name : string
        name of the Excel file to read from

    Returns
    -------
    rankings : list of 1D-numpy integer arrays
        1-based rankings
    stakeholders : list of strings
        names of the stakeholders associated to the rankings
    """
    workbook = oxl.load_workbook(filename=file_name, data_only=True)

    # open the worksheet "Ranking"
    worksheet = workbook['Ranking']

    # maximal row to consider at all (The openpyxl-function worksheet.max_row is sometimes unreliable and may return
    # too large values.)
    max_row = 300

    # maximal column to consider at all (The openpyxl-function worksheet.max_column is sometimes unreliable and may
    # return too large values)
    max_column = 100

    stakeholders = []
    rankings = []

    # loop over the first row to get the number of rankings
    nranks = 0
    for icol in range(1, max_column + 1):
        if worksheet.cell(row=1, column=icol).value is not None:
            stakeholders.append(worksheet.cell(row=1, column=icol).value)
            nranks = nranks + 1
        else:
            break

    for irank in range(0, nranks):
        rank_length = 0
        for irow in range(0, max_row):
            if worksheet.cell(row=irow + 2, column=irank + 1).value is not None:
                rank_length = rank_length + 1
            else:
                break

        rankings.append(np.zeros(rank_length, dtype=int))

        for irow in range(0, rank_length):
            rankings[irank][irow] = worksheet.cell(row=irow + 2, column=irank + 1).value

    return rankings, stakeholders


def read_shares_from_excel(file_name: str) -> Tuple[np.ndarray, list, list]:
    """
    This function reads observed shares from an Excel file.
    It is assumed that the data are contained in a worksheet called "Shares". This sheet title is case-sensitive. The
    first row is reserved for the names of the stakeholders associated with the shares, which are given column-wise
    in Excel. THe first column is reserved for the names of the alternatives.

    Parameters
    ----------
    file_name : string
        name of the Excel file to read from

    Returns
    -------
    shares : list of 1D-numpy float arrays
        shares, i.e. the probability of choosing the alternative
    alternatives : list of strings
        names of the alternatives
    stakeholders : list of strings
        names of the stakeholders associated to the shares
    """

    # the option "data_only=True" ensures that instead of a formular or a reference to another cell, the value
    # displayed in Excel is read in.
    workbook = oxl.load_workbook(filename=file_name, data_only=True)

    # open the worksheet "Shares"
    worksheet = workbook['Shares']

    # maximal row to consider at all (The openpyxl-function worksheet.max_row is sometimes unreliable and may return
    # too large values.)
    max_row = 300

    # maximal column to consider at all (The openpyxl-function worksheet.max_column is sometimes unreliable and may
    # return too large values)
    max_column = 100

    stakeholders = []
    alternatives = []

    eps = 10 * np.finfo(float).eps

    # loop over the first row to get the number of shares, i.e. the number of stakeholders
    nshares = 0
    for icol in range(1, max_column + 1):
        if worksheet.cell(row=1, column=icol + 1).value is not None:
            stakeholders.append(worksheet.cell(row=1, column=icol + 1).value)
            nshares = nshares + 1
        else:
            break

    if nshares == 0:
        err_msg = ('No stakeholders were found in the sheet. Please provide at least one stakeholder with the '
                   'corresponding shares.')
        raise NameError(err_msg)

    # read the names of the alternatives
    for irow in range(1, max_row):
        if worksheet.cell(row=irow + 1, column=1).value is not None:
            alternatives.append(worksheet.cell(row=irow + 1, column=1).value)
        else:
            break

    if len(alternatives) == 0:
        err_msg = 'No alternatives have been provided at all. Please provide at least two alternatives.'
        raise NameError(err_msg)

    if len(alternatives) == 1:
        err_msg = 'Just one alternative has been provided. Please provide at least two alternatives.'
        raise NameError(err_msg)

    nalts = len(alternatives)

    shares = []
    for ishare in range(nshares):
        shares.append(np.zeros(nalts))

    #shares = np.zeros((nalts, nshares), dtype=float)

    for icol in range(0, nshares):
        for irow in range(0, nalts):
            if worksheet.cell(row=irow + 2, column=icol + 1).value is not None:
                shares[icol][irow] = worksheet.cell(row=irow + 2, column=icol + 2).value
            else:
                err_msg = 'Empty cell during reading shares found.'
                raise NameError(err_msg)

    if np.any(shares < -eps):
        col_neg = np.where(shares < -eps)[1][0]
        err_msg = 'Negative shares detected for stakeholder ' + str(stakeholders[col_neg]) + \
                  '. As shares consist of probabilities, they must be non-negative.'
        raise NameError(err_msg)

    if np.any(shares > 1 + eps):
        col_large = np.where(shares > 1 + eps)[1][0]
        err_msg = 'Shares greater than one detected for stakeholder ' + str(stakeholders[col_large]) \
                  + '. As the shares consist of probabilities, they must be at most one.'
        raise NameError(err_msg)

    for ishare in range(0, nshares):
        sum_of_shares = np.sum(shares[ishare])
        if np.abs(sum_of_shares - 1.0) > 10 * eps:
            err_msg = 'The sum of shares for stakeholder ' + str(stakeholders[ishare]) + ' is not one, but ' + \
                      str(sum_of_shares) + '.'
            raise NameError(err_msg)

    return shares, alternatives, stakeholders


def __check_criteria(criteria_key: list, criteria_mean: list, key: str):
    if not criteria_key == criteria_mean:
        if not criteria_mean[0] == criteria_key[0]:
            err_msg = 'The first criterion given for \'Mean\' does not match the one given for \'' + str(key) + '\'.'
            raise NameError(err_msg)

        if len(criteria_mean) > 1:
            if not criteria_mean[1] == criteria_key[1]:
                err_msg = ('The second criterion given for \'Mean\' does not match the one given for \'' + str(key)
                           + '\'.')
                raise NameError(err_msg)

        if len(criteria_mean) > 2:
            if not criteria_mean[2] == criteria_key[2]:
                err_msg = ('The third criterion given for \'Mean\' does not match the one given for \'' + str(key)
                           + '\'.')
                raise NameError(err_msg)

        if len(criteria_mean) > 3:
            for i in range(3, len(criteria_mean)):
                if not criteria_mean[i] == criteria_key[i]:
                    err_msg = 'The ' + str(i + 1) + \
                              'th criterion given for \'Mean\' does not match the one given for \'' + str(key) + '\'.'
                    raise NameError(err_msg)
    return None


def read_bounds_from_excel(file_name: str) -> Tuple[list, list, list]:
    """
    This function reads upper and lower bounds for weights from an Excel file for all stakeholders considered.
    It is assumed that the data are contained in a worksheet called bounds. This sheet title is case-sensitive. The
    first row is reserved for the names of the stakeholders associated with the rankings, which are given column-wise
    in Excel.
    It is assumed in the Excel file, that any of the weights are bounded and that the order of weights corresponds to
    their order of the performance matrix. As the sum of weights is one, providing an upper bound of 1 and a lower
    bound of 0 corresponds to no constraints.
    For additional flexibility, upper and lower bounds are reformulated as linear conditions of the shape
    :math:`Aw \\leq b`, :math:`w` being the weight, as this format allows for arbitrary linear constraints
    like e.g. :math:`w_1 < w_2` in the future.

    Parameters
    ----------
    file_name : string
        name of the Excel file to read the bounds from

    Returns
    -------
    all_a_constraints: list of 2D numpy arrays
        all_a_constraints[i] contains the constraints matrix for the i-th stakeholder
    all_b_constraints: list of 1D numpy arrays
        all_b_constraints[i] contains the constraints right-hand side for the i-th stakeholder
    stakeholders : list of strings
        names of the stakeholders associated to the rankings
    """
    eps = 1e-10
    row_min = 0
    row_max = 0

    workbook = oxl.load_workbook(filename=file_name, data_only=True)

    # read in the characteristics
    worksheet = workbook['bounds']

    # maximal row to consider at all (The openpyxl-function worksheet.max_row is sometimes unreliable and may return
    # too large values.)
    max_row = 300

    # maximal column to consider at all (The openpyxl-function worksheet.max_column is sometimes unreliable and may
    # return too large values)
    max_column = 100

    # find keyword "MINVALUES" in first column
    min_found = False
    for irow in range(1, max_column):
        if worksheet.cell(row=irow, column=1).value == 'MINVALUE':
            row_min = irow
            min_found = True

    # find keyword "MAXVALUES" in first column
    max_found = False
    for irow in range(1, max_column):
        if worksheet.cell(row=irow, column=1).value == 'MAXVALUE':
            row_max = irow
            max_found = True

    if not min_found and not max_found:
        raise NameError('Neither upper nor lower bounds provided')

    offset = row_min + 2

    # get number of criteria by counting the non-empty cells in the first column
    ncrit = 0
    for i in range(offset, max_row):
        if worksheet.cell(column=1, row=i).value is not None:
            ncrit = ncrit + 1
        else:
            break

    all_a_constraints: list[ndarray] = []
    all_b_constraints = []
    stakeholders = []

    # read in the column below the keyword "MINVALUE" (contains the names of
    # the stakeholders) and get the number of stakeholders
    for icolumn in range(2, worksheet.max_row):
        if worksheet.cell(row=2, column=icolumn).value is not None:
            stakeholders.append(worksheet.cell(row=2, column=icolumn).value)

    n_stakeholders = len(stakeholders)

    # read in lower bounds

    # run over columns
    for icolumn in range(2, n_stakeholders + 2):
        n_constrs = 0

        # consistency check
        for irow in range(offset, ncrit + offset):
            irow_max = irow - offset + row_max + 2
            if abs(float(worksheet.cell(row=irow, column=icolumn).value) -
                   float(worksheet.cell(row=irow_max, column=icolumn).value)) < eps:
                err_msg = 'For stakeholder ' + stakeholders[icolumn - 2] + \
                          ', a lower bound equals an upper bound. ' + \
                          'This is not supported yet.'
                raise NameError(err_msg)

            if float(worksheet.cell(row=irow_max, column=icolumn).value) - \
                    float(worksheet.cell(row=irow, column=icolumn).value) < -eps:
                err_msg = 'For stakeholder ' + stakeholders[icolumn - 2] + \
                          ', a lower bound is larger than an upper bound. '
                raise NameError(err_msg)

        for irow in range(offset, ncrit + offset):
            if float(worksheet.cell(row=irow, column=icolumn).value) > 0:
                n_constrs = n_constrs + 1

        a_constraints = np.zeros((n_constrs, ncrit))
        b_constraints = np.zeros(n_constrs)
        iidx = 0
        for irow in range(offset, ncrit + offset):
            if float(worksheet.cell(row=irow, column=icolumn).value) > 0:
                a_constraints[iidx, irow - offset] = -1.0
                b_constraints[iidx] = \
                    -float(worksheet.cell(row=irow, column=icolumn).value)
                iidx = iidx + 1

        all_a_constraints.append(a_constraints)
        all_b_constraints.append(b_constraints)

    offset = row_max + 2

    # read in upper bounds
    for icolumn in range(2, n_stakeholders + 2):
        n_constrs = 0
        for irow in range(offset, ncrit + offset):
            if float(worksheet.cell(row=irow, column=icolumn).value) < 1:
                n_constrs = n_constrs + 1

        a_constraints_max = np.zeros((n_constrs, ncrit))
        b_constraints_max = np.zeros(n_constrs)
        iidx = 0
        for irow in range(offset, ncrit + offset):
            if float(worksheet.cell(row=irow, column=icolumn).value) < 1.0:
                a_constraints_max[iidx, irow - offset] = 1.0
                b_constraints_max[iidx] = \
                    float(worksheet.cell(row=irow, column=icolumn).value)
                iidx = iidx + 1

        all_a_constraints[icolumn - 2] = \
            np.vstack((all_a_constraints[icolumn - 2], a_constraints_max))

        all_b_constraints[icolumn - 2] = \
            np.hstack((all_b_constraints[icolumn - 2], b_constraints_max))

    return all_a_constraints, all_b_constraints, stakeholders


def read_weights_from_excel(file_name: str) -> Tuple[list, list, list]:
    """
    This function reads the weights of criteria from an Excel file. It is assumed that the data are contained in a
    worksheet called "Weights". This sheet title is case-sensitive. The first row is reserved for the names of the
    stakeholders associated with the weights, which are given column-wise in Excel. The weights are expected to start
    in the first column.

    Parameters
    ----------
    file_name : string
        name of the Excel file to read from

    Returns
    -------
    criteria : list of strings
        names of the criteria
    weights : list of 1D-numpy integer arrays
        1-based rankings
    stakeholders : list of strings
        names of the stakeholders associated to the weights
    """
    workbook = oxl.load_workbook(filename=file_name, data_only=True)

    # open the worksheet "Ranking"
    worksheet = workbook['Weights']

    # maximal row and column index occurring in this shet
    # maximal row to consider at all (The openpyxl-function worksheet.max_row is sometimes unreliable and may return
    # too large values.)
    max_row = 300

    # maximal column to consider at all (The openpyxl-function worksheet.max_column is sometimes unreliable and may
    # return too large values)
    max_column = 100

    criteria = []
    stakeholders = []
    weights = []

    # loop over the first row to get the number of stakeholders and their names
    nstakeholders = 0
    for icol in range(2, max_column + 1):
        if worksheet.cell(row=1, column=icol).value is not None:
            stakeholders.append(worksheet.cell(row=1, column=icol).value)
            nstakeholders = nstakeholders + 1
        else:
            break

    # loop over the first column to get the number of criteria and their names
    ncrit = 0
    for irow in range(0, max_row):
        if worksheet.cell(row=irow + 2, column=1).value is not None:
            criteria.append(worksheet.cell(row=irow + 2, column=1).value)
            ncrit = ncrit + 1
        else:
            break

    for istakeholder in range(0, nstakeholders):
        weights.append(np.zeros(ncrit, dtype=float))

        for irow in range(0, ncrit):
            aux_val = worksheet.cell(row=irow + 2, column=istakeholder + 2).value
            if aux_val is not None:
                if aux_val > -1e-12:
                    weights[istakeholder][irow] = aux_val
                else:
                    err_msg = 'There are negative weights for stakeholder ' + stakeholders[istakeholder] + '. Only ' + \
                              'non-negative weights are allowed.'
                    raise NameError(err_msg)
            else:
                err_msg = 'Reading the weights for stakeholder ' + stakeholders[istakeholder] + ' failed due ' + \
                          'to an empty cell. Please provide any weight.'
                raise NameError(err_msg)

    # test, if all weighting vectors are normalized
    for istakeholder in range(0, nstakeholders):

        if abs(np.sum(weights[istakeholder]) - 1.0) > 0.00001:
            warn_msg = 'The weights for stakeholder ' + stakeholders[istakeholder] + ' have been scaled such that' + \
                ' their sum is one.'
            warnings.warn(warn_msg)

            weights[istakeholder] = weights[istakeholder]/np.sum(weights[istakeholder])

    return criteria, stakeholders, weights


def read_params_from_excel(file_name: str, sheet_name: str):
    """
    This function reads the MCDA-model-related parameters from the sheet sheet_name in the Excel file named file_name.
    This sheet title is case-sensitive. The keywords are expected to be located in the first column, the data associated
    to the keywords are expected to be located directly below the keywords.

    Parameters
    ----------
    file_name : string
        name of the Excel file to read from
    sheet_name : string
        name of the sheet in the Excel file to read from

    Returns
    -------
    prom2 : boolean
        if True, the MCDA model is Promethee 2, otherwise SAW
    params: ndarray of two ndarrays; params[0]: type of preference function (integer), params[1]: empty array for SAW,
        array of parameters of the preferences functions for Promethee 2
    """
    workbook = oxl.load_workbook(filename=file_name, data_only=True)

    # open the worksheet
    worksheet = workbook[sheet_name]

    # maximal row and column index occurring in sheet named sheet_name
    # maximal row and column index occurring in this shet
    # maximal row to consider at all (The openpyxl-function worksheet.max_row is sometimes unreliable and may return
    # too large values.)
    max_row = 300

    # maximal column to consider at all (The openpyxl-function worksheet.max_column is sometimes unreliable and may
    # return too large values)
    max_column = 100

    # search the keyword "Method" in the first column and read the content of the cell directly below.
    prom2 = False
    for irow in range(1, max_row + 1):
        if worksheet.cell(row=irow, column=1).value == 'Method':
            prom2 = worksheet.cell(row=irow + 1, column=1).value != 'SAW'
            break

    type_criterion = []
    params_prom2 = [[], []]

    # read MCDA-parameters in the case of Promethee 2 (there are no parameters for SAW).
    if prom2:

        # Search keyword 'Types' in the first column and start reading directly below for creating a list of types of
        # the preference functions.
        for irow in range(1, max_row + 1):
            if worksheet.cell(row=irow, column=1).value == 'Types':
                for icol in range(1, max_column + 1):
                    value = worksheet.cell(row=irow + 1, column=icol).value
                    if value is not None:
                        type_criterion += [value]
                    else:
                        break
                break

        # Same for the parameters of the preference functions
        for irow in range(1, max_row + 1):
            if worksheet.cell(row=irow, column=1).value == 'Parameters':
                for offset in range(1, 3):
                    for icol in range(1, max_column + 1):
                        value = worksheet.cell(row=irow + offset, column=icol).value
                        if value is not None:
                            params_prom2[offset - 1] += [value]
                        else:
                            break
                break

    return prom2, np.array(type_criterion, dtype=int), np.array(params_prom2, dtype=float)


def main():
    file_name = '../examples/input_simple_analysis.xlsx'
    perfmat_mean, perfmat_type, perfmat_param, benefit, criteria, alternatives = \
        read_perfmat_from_excel(file_name, 'Characteristics')

    print('mean values:')
    print(perfmat_mean)

    print()
    print('type of matrix entries:')
    print(perfmat_type)

    print()
    print('cost/benefit:')
    print(benefit)

    rankings, stakeholders = read_rankings_from_excel(file_name)
    print()
    print('stakeholders:', stakeholders)

    file_name = '../examples/input_bayesian_analysis.xlsx'
    rankings, alternatives, stakeholders = read_shares_from_excel(file_name)


if __name__ == '__main__':
    main()
