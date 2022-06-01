# -*- coding: utf-8 -*-
"""
Created on Sat Feb 13 19:16:06 2021

This module provides a simple example for computing weights from rankings with
pysmaa.

@author: Matthias Grajewski, FH Aachen University of Applied Sciences
"""
import sys
from os import path
import numpy as np
import openpyxl as oxl

# ---------------------- change this to run own analysis ----------------------
# json-File with all necessary data and parameters for computation
COMP_PARS_FILE = 'simple_analysis'
# ----------------------------------------------------------------------------

# if this module is executed stand-alone,
# we want to import pysmaa without hard-coding its location here. Instead, we
# provide the path to pysmaa\src in the json-control-file and read it
# here.
if __name__ == '__main__':
    import json as js

    with open(f'{COMP_PARS_FILE}.json') as file:
        parameter_dict = js.load(file)

    path_to_pysmaa = parameter_dict['path_to_pysmaa']

    # if a path to pysmaa is provided, read it and append if, if not
    # already present in sys.path
    if path_to_pysmaa != '':
        if not any(path.normcase(sp) == path_to_pysmaa for sp in sys.path):
            sys.path.append(f'{path_to_pysmaa}/src')

# now, we can import all the functions from pysmaa
from weights_from_rankings import p_from_ptilde_saw, p_from_ptilde_prom_2, \
                                  inscribe_maximal_ellipsoid, check_points, \
                                  hrep_from_ranking, compute_margins, get_q

from read_from_Excel import read_ptilde_from_excel, read_rankings_from_excel, \
                            read_bounds_from_excel
from inoutput import read_comp_pars


def perform_simple_analysis(comp_pars_file):
    """
    This function is a wrapper for the actual computations of the weights from the rankings and the associated
    quantities. The names of the Excel input files (this can be more than one) are provided in `comp_pars_file`.
    The rankings (one-based!) are provided by an Excel file. The routine for reading the rankings expects a worksheet
    named 'Ranking' within this file.
    It is furthermore assumed that the names of the rankings are provided in the first row of the sheet and that the
    rankings are given as column vectors starting from the second row. The indices of the alternatives are meant to be
    sorted in descending order.
    The same Excel file must contain a worksheet named 'Characteristics' in which the performance matrix is stored. One
    column represents one criterion for all alternatives, such that the k-th row of the matrix represents all values of
    the k-th alternative. The first column and the first row are reserved for the names of the criteria and
    alternatives, such that the performance matrix must start in the second column and row in the Excel worksheet.
    This functions writes an Excel file with the results; its name is provided as well in `comp_pars_file`.

    Parameters
    ----------
    comp_pars_file : json-file
        It contains all necessary information for running the analyses. For details, we refer to the documentation
        provided in inoutput.py

    Returns
    -------
    None
    """

    # dummy initializations
    a_constraints = 0
    b_constraints = 0
    U = 0
    offset = 0

    # read all computational parameters from the json-file
    this_pars = read_comp_pars(comp_pars_file)

    # read the performance matrix
    ptilde = read_ptilde_from_excel(this_pars.input_file_P)[0]

    # read the rankings and the names of the corresponding milieus
    rankings, milieus = read_rankings_from_excel(this_pars.input_file_ranks)

    if this_pars.input_file_bounds != '':
        a_constraints, b_constraints, milieus = read_bounds_from_excel(this_pars.input_file_bounds)

    ncrit = ptilde.shape[1]

    # choose preference function type in Promethee II
    type_criterion = this_pars.promethee_pref_func_type*np.ones(ncrit, dtype=int)
    benefit = np.full(ncrit, True)

    mcda_type = this_pars.mcda_type

    # create Excel workbook for the results of our computation
    results_in_excel = oxl.Workbook()

    # loop over all given rankings and analyse the corresponding polytopes W_r
    for iranking in range(0, len(rankings)):

        print('')
        print('----- analyzing ranking for milieu ', milieus[iranking], '-----')

        # rankings are 0-based from here on!
        current_ranking = rankings[iranking] - 1

        # compute performance matrix according to the MCDA model
        if mcda_type == 'Promethee':

            # setup of the parameter vector
            parameters = np.zeros((2, ncrit))
            parameters[0, :] = this_pars.promethee_c
            p = p_from_ptilde_prom_2(ptilde, type_criterion, parameters, benefit)
        elif mcda_type == 'SAW':
            # compute performance matrix according to SAW
            p = p_from_ptilde_saw(ptilde)
        else:
            raise NameError('Invalid choice of MCDA model.')

        # analyse the W_r
        if this_pars.input_file_bounds != '':
            w_el, b_el = inscribe_maximal_ellipsoid(p, current_ranking, a_constraints[iranking],
                                                    b_constraints[iranking])
        else:
            w_el, b_el = inscribe_maximal_ellipsoid(p, current_ranking)

        # start writing to Excel
        results_in_excel.create_sheet(milieus[iranking])
        current_sheet = results_in_excel[milieus[iranking]]

        # consistency checks
        if w_el.shape[0] > 0:
            check = check_points(p, current_ranking, np.array([w_el]))
            print('check w inside W_r:', check[0])

        if w_el.shape[0] > 0:
            U, sizes = np.linalg.svd(b_el)[0:2]
            sizes = 2*sizes
        else:
            sizes = np.zeros(ncrit)

        current_sheet.cell(row=1, column=1).value = milieus[iranking]
        current_sheet.cell(row=1, column=1).font = oxl.styles.Font(size=14, bold=True)

        current_sheet.cell(row=2, column=1).value = 'ranking'

        for i in range(0, current_ranking.shape[0]):
            # we want the ranking 1-based!
            current_sheet.cell(row=2, column=i+2).value = current_ranking[i]+1

        current_sheet.cell(row=4, column=1).value = 'weights'

        if w_el.shape[0] > 0:
            for i in range(0, w_el.shape[0]):
                current_sheet.cell(row=5, column=i+2).value = w_el[i]
                current_sheet.cell(row=5, column=i+2).number_format = '0.00'

            current_sheet.cell(row=7, column=1).value = 'main axes (column vectors)'
            for i in range(0, U.shape[0]-1):
                for j in range(0, U.shape[1]):
                    # in sigma_aux, the ROWS form the principal directions, however, we
                    # want them to be COLUMN vectors
                    current_sheet.cell(row=8+j, column=i+2).value = U[j, i]
                    current_sheet.cell(row=8+j, column=i+2).number_format = '0.00'

            offset = 8 + U.shape[1]

        if w_el.shape[0] > 0:

            current_sheet.cell(row=offset + 1, column=1).value = 'size of inscribed ellipsoid'

            for i in range(0, sizes.shape[0]):
                current_sheet.cell(row=offset+2, column=i+2).value = sizes[i]

        offset = offset + 4
        current_sheet.cell(row=offset + 1, column=1).value = 'margins aka distance to hyperplanes'

        # margins apply to the actual decision planes only, therefore we do not
        # incorporate the user-prescribed constraints
        A, b = hrep_from_ranking(p, current_ranking)

        if w_el.shape[0] > 0:

            # shift the polytope by the center of gravity of the standard simplex
            cog = 1/ncrit*np.ones(ncrit)
            bhat = b - A@cog
            # rotate it such that the last coordinate is zero
            q = get_q(ncrit)
            Ahat = A@q
            wel_hat = q.T@(w_el - cog)

            # the very last two margins are zero anyway, so we do not need to
            # compute them. Moreover, we discard the last coordinate
            # (zero anyway).
            margins = compute_margins(Ahat[:-2, :-1], bhat[:-2], wel_hat[:-1])

            margins_others = min(margins[ncrit+current_ranking.shape[0]-1:])

            # consider only the margins concerning the actual ranking
            margins = margins[ncrit:ncrit+current_ranking.shape[0]-1]

            for i in range(margins.shape[0]):
                current_sheet.cell(row=offset+2, column=i+2).value = \
                    str(current_ranking[i]+1) + ' > ' + str(current_ranking[i+1]+1)
                current_sheet.cell(row=offset+3, column=i+2).value = margins[i]

            current_sheet.cell(row=offset+2, column=margins.shape[0]+2).value \
                = str(current_ranking[-1]+1) + ' > others'
            current_sheet.cell(row=offset+3, column=margins.shape[0]+2).value = margins_others

    print()
    print('write results to Excel file', this_pars.output_file)

    # Save results in Excel-File
    results_in_excel.save(this_pars.output_file)


# --------------------------------------------------------------------------
# x
# x This is the actual function call.
# --------------------------------------------------------------------------

if __name__ == '__main__':
    perform_simple_analysis(COMP_PARS_FILE)
